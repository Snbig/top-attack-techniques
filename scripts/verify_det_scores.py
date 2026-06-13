"""Verify the detection_score distribution in the new Calculator.xlsx."""
from openpyxl import load_workbook
from collections import Counter

wb = load_workbook(r'C:\Users\Administrator\AppData\Local\Temp\top-attack\src\data\Calculator.xlsx', data_only=True)
ws = wb['Methodology']

scores = Counter()
none_count = 0
total = 0

for row in ws.iter_rows(min_row=2, values_only=False):
    tid_cell = row[2]
    if tid_cell and tid_cell.value:
        tid = str(tid_cell.value).strip()
        ds = row[27].value  # Column AB
        cs = row[21].value  # Column V
        mit = row[24].value  # Column Y
        total += 1
        
        if ds is None:
            none_count += 1
        else:
            rounded = round(float(ds), 2)
            scores[rounded] += 1

wb.close()

print(f'Total techniques in Methodology: {total}')
print(f'None detection_score: {none_count}')
print(f'\nDetection score distribution (rounded to 0.01):')
flat_01 = 0
for score in sorted(scores.keys()):
    count = scores[score]
    marker = ' *** FLAT' if abs(score - 0.01) < 0.005 else ''
    if abs(score - 0.01) < 0.005:
        flat_01 += count
    bar = '#' * min(count, 50)
    print(f'  {score:.2f}: {count:4d} {bar}')
    
print(f'\nTotal flat 0.01: {flat_01}')
print(f'Total unique values: {len(scores)}')
print(f'Scores cover range: {min(scores.keys()):.2f} - {max(scores.keys()):.2f}')
