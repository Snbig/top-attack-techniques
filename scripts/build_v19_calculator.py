"""
Build a v19.1 Calculator.xlsx from STIX data + preserve existing Methodology scoring.
Includes CTID methodology scoring for new techniques using attackcti.

Usage: python scripts/build_v19_calculator.py
"""
import json
import math
import os
import sys
from copy import copy
from collections import defaultdict, Counter

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from attackcti import attack_client

REPO_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
STIX_PATH = os.path.join(REPO_DIR, "enterprise-attack-v19.1.json")
OLD_CALC_PATH = os.path.join(REPO_DIR, "src", "data", "Calculator.xlsx")
NEW_CALC_PATH = os.path.join(REPO_DIR, "src", "data", "Calculator.xlsx")
EXTERNAL_COUNTS_PATH = os.path.join(REPO_DIR, "scripts", "external_detection_counts.json")

# === REVOCATION MAP: v14.1 -> v19.1 ===
# Techniques that were revoked and replaced
REVOCATION_MAP = {
    "T1562":     "T1685",      # Impair Defenses -> Disable or Modify Tools
    "T1562.001": "T1685",      # Disable or Modify Tools -> parent
    "T1562.002": "T1685.001",  # Disable Windows Event Logging -> Disable or Modify Windows Event Log
    "T1562.003": "T1690",      # Impair Command History Logging -> Prevent Command History Logging
    "T1562.004": "T1686",      # Disable or Modify System Firewall -> Disable or Modify System Firewall (new parent)
    "T1562.005": "T1688",      # Safe Mode Boot -> Safe Mode Boot (new parent)
    "T1562.006": "T1685",      # Indicator Blocking -> parent
    "T1562.007": "T1686.001",  # Disable or Modify Cloud Firewall -> Cloud Firewall
    "T1562.008": "T1685.002",  # Disable or Modify Cloud Logs -> Disable or Modify Cloud Log
    "T1562.009": "T1685.004",  # Disable or Modify Linux Audit System -> Disable or Modify Linux Audit System Log
    "T1562.010": "T1689",      # Downgrade Attack -> Downgrade Attack (new parent)
    "T1562.011": "T1685.003",  # Spoof Security Alerting -> Modify or Spoof Tool UI
    "T1562.012": "T1690",      # from the v19 changes: Impair Command History Logging -> Prevent Command History Logging
    "T1070.001": "T1685.005",  # Clear Windows Event Logs -> Clear Windows Event Logs
    "T1070.002": "T1685.006",  # Clear Linux or Mac System Logs -> Clear Linux or Mac System Logs
    "T1574.002": "T1574.001",  # DLL Side-Loading -> DLL (was DLL Search Order Hijacking)
    "T1656":     "T1684.001",  # Impersonation -> Social Engineering: Impersonation
    "T1672":     "T1684.002",  # Email Spoofing -> Social Engineering: Email Spoofing
}

# Renamed techniques
RENAME_MAP = {
    "T1574.001": "DLL",  # was "DLL Search Order Hijacking"
    "T1554": "Compromise Host Software Binary",  # was "Compromise Client Software Binary"
    "T1484": "Domain or Tenant Policy Modification",  # was "Domain Policy Modification"
    "T1211": "Exploitation for Stealth",  # was "Exploitation for Defense Evasion"
}

def load_stix():
    """Load STIX v19.1 data."""
    with open(STIX_PATH, encoding="utf-8") as f:
        return json.load(f)

def get_external_id(obj, prefix=None):
    """Get external_id from STIX object, optionally filtering by prefix."""
    for ref in obj.get("external_references", []):
        eid = ref.get("external_id", "")
        if prefix:
            if eid.startswith(prefix):
                return eid, ref.get("url", "")
        elif eid:
            return eid, ref.get("url", "")
    return None, None

def extract_techniques(stix):
    """Extract all enterprise attack patterns (techniques)."""
    tid_map = {}  # TID -> technique info
    id_to_tid = {}  # STIX ID -> TID
    
    for obj in stix["objects"]:
        if obj.get("type") != "attack-pattern":
            continue
        # Skip non-enterprise
        if "enterprise-attack" not in obj.get("x_mitre_domains", []):
            continue
        
        tid, url = get_external_id(obj, prefix="T")
        if not tid:
            continue
        
        name = obj.get("name", "")
        description = obj.get("description", "")
        is_sub = obj.get("x_mitre_is_subtechnique", False)
        platforms = obj.get("x_mitre_platforms", [])
        revoked = obj.get("revoked", False)
        deprecated = obj.get("x_mitre_deprecated", False)
        data_sources = obj.get("x_mitre_data_sources", [])
        detection = obj.get("x_mitre_detection", "")
        
        # Get tactic info from kill_chain_phases
        tactics = []
        for kcp in obj.get("kill_chain_phases", []):
            if kcp.get("kill_chain_name") == "mitre-attack":
                tactics.append(kcp.get("phase_name", ""))
        
        info = {
            "tid": tid,
            "name": name,
            "description": description,
            "url": url,
            "platforms": platforms,
            "is_subtechnique": is_sub,
            "revoked": revoked,
            "deprecated": deprecated,
            "tactics": tactics,
            "data_sources": data_sources,
            "detection": detection,
            "subtechniques": [],
            "mitigations": [],
        }
        
        tid_map[tid] = info
        id_to_tid[obj["id"]] = tid
    
    # Determine parent for sub-techniques
    for obj in stix["objects"]:
        if obj.get("type") != "attack-pattern":
            continue
        if "enterprise-attack" not in obj.get("x_mitre_domains", []):
            continue
        
        tid, _ = get_external_id(obj, prefix="T")
        if not tid or "." not in tid:
            continue
        
        parent_tid = tid.split(".")[0]
        if parent_tid in tid_map:
            tid_map[parent_tid]["subtechniques"].append(tid)
    
    return tid_map, id_to_tid

def extract_mitigations(stix):
    """Extract mitigations (course-of-action with M-prefix)."""
    mitigations = {}
    for obj in stix["objects"]:
        if obj.get("type") != "course-of-action":
            continue
        mid, url = get_external_id(obj, prefix="M")
        if not mid:
            continue
        
        mitigations[mid] = {
            "mid": mid,
            "name": obj.get("name", ""),
            "description": obj.get("description", ""),
            "url": url,
            "deprecated": obj.get("x_mitre_deprecated", False),
        }
    return mitigations

def extract_relationships(stix, id_to_tid, mitigations):
    """Extract mitigation-to-technique relationships."""
    technique_mitigations = defaultdict(list)
    
    for obj in stix["objects"]:
        if obj.get("type") != "relationship":
            continue
        if obj.get("relationship_type") != "mitigates":
            continue
        
        source_ref = obj.get("source_ref", "")
        target_ref = obj.get("target_ref", "")
        
        # We want course-of-action -> attack-pattern
        target_tid = id_to_tid.get(target_ref)
        source_mid = None
        
        # Find MID for the source course-of-action
        for src_obj in stix["objects"]:
            if src_obj.get("id") == source_ref and src_obj.get("type") == "course-of-action":
                source_mid, _ = get_external_id(src_obj, prefix="M")
                break
        
        if source_mid and target_tid and source_mid in mitigations:
            technique_mitigations[target_tid].append(source_mid)
    
    return technique_mitigations

def load_existing_methodology():
    """Load existing Calculator.xlsx and extract Methodology sheet data."""
    if not os.path.exists(OLD_CALC_PATH):
        print("No existing Calculator.xlsx found. Starting fresh.")
        return {}, {}, {}, {}
    
    wb = load_workbook(OLD_CALC_PATH, data_only=True)
    
    # Extract Methodology sheet
    methodology = {}
    if "Methodology" in wb.sheetnames:
        ws = wb["Methodology"]
        for row in ws.iter_rows(min_row=2, values_only=False):
            tid_cell = row[2]  # Column C (index 2)
            if tid_cell and tid_cell.value:
                tid = str(tid_cell.value).strip()
                methodology[tid] = {
                    "cumulative_score": row[1].value,  # B
                    "choke_point_score": row[7].value,  # H (index 7)
                    "prevalence_score": row[12].value,  # M (index 12)
                    "has_car": bool(row[13].value),  # N
                    "has_sigma": bool(row[14].value),  # O
                    "has_es_siem": bool(row[15].value),  # P
                    "has_splunk": bool(row[16].value),  # Q
                    "cis_controls": str(row[17].value or ""),  # R
                    "nist_controls": str(row[19].value or ""),  # T
                    "process_coverage": bool(row[30].value),  # AE (index 30)
                    "network_coverage": bool(row[32].value),  # AG
                    "file_coverage": bool(row[34].value),  # AI
                    "cloud_coverage": bool(row[36].value),  # AK
                    "hardware_coverage": bool(row[38].value),  # AM
                    "combined_score": row[21].value,  # V (index 21)
                    "mitigation_score": row[24].value,  # Y
                    "detection_score": row[27].value,  # AB
                }
    
    wb.close()
    print(f"Loaded Methodology data for {len(methodology)} techniques from existing Calculator.xlsx")
    return methodology

def load_external_counts():
    """Load external_detection_counts.json (CAR/Sigma/ES/Splunk + CTID sightings)."""
    if not os.path.exists(EXTERNAL_COUNTS_PATH):
        print(f"  WARNING: {EXTERNAL_COUNTS_PATH} not found. Running without external counts.")
        return {}
    with open(EXTERNAL_COUNTS_PATH, encoding="utf-8") as f:
        data = json.load(f)
    print(f"Loaded external counts for {len(data)} techniques")
    with_det = sum(1 for v in data.values() if v.get("total_detections", 0) > 0)
    with_sight = sum(1 for v in data.values() if v.get("sightings", 0) > 0)
    print(f"  With detection rules: {with_det}, With CTID sightings: {with_sight}")
    return data

# ===== CTID METHODOLOGY SCORING =====
# Formula: A(x_d, x_m) = w_d * u_d(x_d) + w_m * u_m(x_m)
# u(x) = piecewise linear: 0 if x<lower, (x-lower)/(upper-lower) if lower<=x<=upper, 1 if x>upper
DET_LOWER, DET_UPPER = 0, 100
MIT_LOWER, MIT_UPPER = 0, 55

# ===== NDI (Normalized Detection Index) =====
# NDI = Σ(W_i × D_i) / Σ(W_i × D_max) × 100
# W_i = technique weight based on tactic (1=low, 2=medium, 3=critical chokepoint)
# D_i = detection level (0=no logs, 1=raw logs, 2=SIEM alert, 3=prevention)
# D_max = 3
# Weight mapping based on MITRE CTID choke point methodology
TACTIC_WEIGHTS = {
    "reconnaissance": 1,        # Low — discovery/recon
    "resource-development": 1,  # Low — pre-attack
    "discovery": 1,             # Low — discovery
    "collection": 2,            # Medium — data access
    "initial-access": 3,        # Critical — chokepoint
    "execution": 3,             # Critical — chokepoint
    "persistence": 3,           # Critical — chokepoint
    "privilege-escalation": 3,  # Critical — chokepoint
    "defense-evasion": 3,       # Critical — chokepoint
    "credential-access": 3,     # Critical — chokepoint
    "lateral-movement": 3,      # Critical — chokepoint
    "command-and-control": 3,   # Critical — chokepoint
    "exfiltration": 3,          # Critical — chokepoint
    "impact": 3,                # Critical — chokepoint
}
D_MAX = 3  # Maximum detection level for NDI normalization

def _utility(x, lower, upper):
    if x <= lower:
        return 0.0
    elif x >= upper:
        return 1.0
    return (x - lower) / (upper - lower)

def compute_weight(tactics):
    """Compute technique weight (1/2/3) based on tactic mapping.
    
    Uses the highest weight among all tactics a technique belongs to.
    """
    if not tactics:
        return 2  # Default to medium if no tactics found
    weights = [TACTIC_WEIGHTS.get(t.lower(), 3) for t in tactics]
    return max(weights)  # Use highest weight among all tactics

def compute_detection_level(tid, external_counts=None, detection_score=0.0, stix_det_count=0):
    """Compute detection level (0-3) based on external detection data.
    
    0 = No logs / no detection capability
    1 = Raw logs only (detection data sources exist but no specific rules)
    2 = SIEM alert (detection rules exist: CAR/Sigma/ES/Splunk)
    3 = Prevention/blocking (inferred from multiple detection rule types)
    """
    if external_counts:
        ext = external_counts.get(tid, {})
        total_det = ext.get("total_detections", 0)
        sightings = ext.get("sightings", 0)
        
        if total_det >= 5:
            # Multiple detection rule types — likely prevention capability
            return 3
        elif total_det > 0:
            # Has detection rules — SIEM alert level
            return 2
        elif sightings > 0:
            # Raw sightings data — forensics/context only
            return 1
    
    # Fallback: STIX detects count or detection_score
    if stix_det_count > 0 or (detection_score and detection_score > 0):
        return 1
    
    return 0

def compute_ndi_scores(techniques_with_w_d):
    """Compute overall NDI for a group of techniques.
    
    NDI = Σ(W_i × D_i) / Σ(W_i × D_max) × 100
    
    Returns (ndi_percentage, numerator_sum, denominator_sum).
    """
    numerator = sum(w * d for w, d in techniques_with_w_d)
    denominator = sum(w * D_MAX for w, _ in techniques_with_w_d)
    if denominator == 0:
        return 0.0, 0, 0
    ndi = (numerator / denominator) * 100
    return round(ndi, 2), numerator, denominator

def compute_actionability(det_count, mit_count):
    """Compute CTID actionability combined score from detection/mitigation counts."""
    w_m = 1.0
    w_d = 0.5 * (DET_UPPER - DET_LOWER) / (MIT_UPPER - MIT_LOWER)
    total = w_d + w_m
    w_d_norm = w_d / total
    w_m_norm = w_m / total
    det_score = _utility(det_count, DET_LOWER, DET_UPPER)
    mit_score = _utility(mit_count, MIT_LOWER, MIT_UPPER)
    combined = w_d_norm * det_score + w_m_norm * mit_score
    return combined, det_score, mit_score

def compute_detection_score_from_external(tid, external_counts, stix_det_count=0):
    """Compute detection_score using external counts (CAR/Sigma/ES/Splunk + CTID sightings).

    Differentiation strategy:
    1. Detection rule counts (total_detections) — sqrt-scaled for granularity at low counts
       sqrt(1)/10=0.10, sqrt(4)/10=0.20, sqrt(9)/10=0.30, sqrt(100)/10=1.0
    2. CTID sightings (log-scaled) — for techniques with 0 detection rules
    3. STIX detects relationship count — linear fallback
    """
    ext = external_counts.get(tid, {})
    total_det = ext.get("total_detections", 0)
    sightings = ext.get("sightings", 0)

    if total_det > 0:
        # Sqrt scale: provides fine granularity at low rule counts
        # 1→0.10, 2→0.14, 3→0.17, 4→0.20, 9→0.30, 25→0.50, 100→1.0
        return min(1.0, math.sqrt(total_det) / 10.0)
    elif sightings > 0:
        # Log scale for sightings (1 to 200K range)
        return min(1.0, math.log10(sightings + 1) / 5.0)
    else:
        # Fallback: STIX detects (also sqrt-scaled for differentiation)
        return min(1.0, math.sqrt(stix_det_count) / 10.0) if stix_det_count > 0 else 0.0

def count_detections(stix, id_to_tid):
    """Count detection strategies and analytics per technique from STIX."""
    detection_counts = Counter()
    analytic_counts = Counter()
    ds_to_technique = {}
    objects_by_id = {o["id"]: o for o in stix["objects"]}
    
    for obj in stix["objects"]:
        if obj.get("type") != "relationship":
            continue
        if obj.get("relationship_type") != "detects":
            continue
        target_tid = id_to_tid.get(obj.get("target_ref", ""))
        source_ref = obj.get("source_ref", "")
        if target_tid and source_ref:
            detection_counts[target_tid] += 1
            ds_to_technique[source_ref] = target_tid
    
    for ds_id, tech_tid in ds_to_technique.items():
        ds_obj = objects_by_id.get(ds_id, {})
        analytic_refs = ds_obj.get("x_mitre_analytic_refs", [])
        analytic_counts[tech_tid] += len(analytic_refs)
    
    return detection_counts, analytic_counts

def build_mid_control_mapping(old_methodology, tech_mitigations):
    """Build MID -> set(CIS_controls) and MID -> set(NIST_controls) mapping."""
    mid_to_cis = defaultdict(set)
    mid_to_nist = defaultdict(set)
    reverse_tech_mit = defaultdict(list)
    
    for tid, mids in tech_mitigations.items():
        for mid in mids:
            reverse_tech_mit[mid].append(tid)
    
    for tid, scores in old_methodology.items():
        if scores.get("combined_score") is not None and scores["combined_score"] != "":
            cis_list = [c.strip() for c in str(scores.get("cis_controls", "")).split(",") if c.strip()]
            nist_list = [n.strip() for n in str(scores.get("nist_controls", "")).split(",") if n.strip()]
            for mid in tech_mitigations.get(tid, []):
                mid_to_cis[mid].update(cis_list)
                mid_to_nist[mid].update(nist_list)
    
    return mid_to_cis, mid_to_nist

def compute_ctid_defaults(tid, tid_map, stix, id_to_tid, detection_counts, 
                           tech_mitigations, mid_to_cis, mid_to_nist, old_methodology,
                           external_counts=None):
    """Compute CTID scores for a new technique that has no existing scores."""
    mit_count = len(tech_mitigations.get(tid, []))
    # Use external counts for detection score if available
    if external_counts:
        det_score = compute_detection_score_from_external(tid, external_counts,
                                                          stix_det_count=detection_counts.get(tid, 0))
    else:
        det_score = _utility(detection_counts.get(tid, 0), DET_LOWER, DET_UPPER)
    mit_score = _utility(mit_count, MIT_LOWER, MIT_UPPER)
    combined, _, _ = compute_actionability(0, mit_count)
    # Recompute combined with our real det_score while preserving mit_score
    w_m = 1.0
    w_d = 0.5 * (DET_UPPER - DET_LOWER) / (MIT_UPPER - MIT_LOWER)
    total = w_d + w_m
    w_d_norm = w_d / total
    w_m_norm = w_m / total
    combined = w_d_norm * det_score + w_m_norm * mit_score
    
    # CIS/NIST from mapped mitigations
    cis_set = set()
    nist_set = set()
    for mid in tech_mitigations.get(tid, []):
        cis_set.update(mid_to_cis.get(mid, set()))
        nist_set.update(mid_to_nist.get(mid, set()))
    
    # Fallback: inherit from parent
    if not cis_set and "." in tid:
        parent_tid = tid.split(".")[0]
        parent_entry = old_methodology.get(parent_tid, {})
        if parent_entry.get("combined_score") is not None:
            cis_set = set(c.strip() for c in str(parent_entry.get("cis_controls", "")).split(",") if c.strip())
            nist_set = set(c.strip() for c in str(parent_entry.get("nist_controls", "")).split(",") if c.strip())
    
    # Fallback: same-tactic techniques
    if not cis_set:
        my_tactics = set()
        stix_obj = next((o for o in stix["objects"] if o["id"] == id_to_tid.get(tid, "")), {})
        for kcp in stix_obj.get("kill_chain_phases", []):
            if kcp.get("kill_chain_name") == "mitre-attack":
                my_tactics.add(kcp.get("phase_name", ""))
        
        for other_tid, other_entry in old_methodology.items():
            if other_entry.get("combined_score") is not None:
                other_obj = next((o for o in stix["objects"] if o["id"] == id_to_tid.get(other_tid, "")), {})
                other_tactics = set()
                for kcp in other_obj.get("kill_chain_phases", []):
                    if kcp.get("kill_chain_name") == "mitre-attack":
                        other_tactics.add(kcp.get("phase_name", ""))
                if my_tactics & other_tactics:
                    cis_set.update(c.strip() for c in str(other_entry.get("cis_controls", "")).split(",") if c.strip())
                    nist_set.update(c.strip() for c in str(other_entry.get("nist_controls", "")).split(",") if c.strip())
    
    # Heuristic choke point (relationship density)
    rel_count = len(tech_mitigations.get(tid, [])) + detection_counts.get(tid, 0)
    cp_score = min(1.0, rel_count / 25.0)
    
    # Heuristic prevalence (software/groups using technique)
    uses_count = 0
    for obj in stix["objects"]:
        if obj.get("type") == "relationship" and obj.get("relationship_type") == "uses":
            if id_to_tid.get(obj.get("target_ref", "")) == tid:
                uses_count += 1
    prev_score = min(1.0, uses_count / 500.0)
    
    # Derive has_* booleans from external counts
    ext = (external_counts or {}).get(tid, {})
    has_car = ext.get("car", 0) > 0
    has_sigma = ext.get("sigma", 0) > 0
    has_es_siem = ext.get("es", 0) > 0
    has_splunk = ext.get("splunk", 0) > 0

    # Compute NDI fields: weight (W), detection_level (D), ndi_score (W×D)
    tactics = tid_map.get(tid, {}).get("tactics", [])
    weight = compute_weight(tactics)
    detection_level = compute_detection_level(tid, external_counts, det_score, detection_counts.get(tid, 0))
    ndi_score = weight * detection_level

    return {
        "cumulative_score": round(combined, 4),
        "choke_point_score": round(cp_score, 4),
        "prevalence_score": round(prev_score, 4),
        "has_car": has_car,
        "has_sigma": has_sigma,
        "has_es_siem": has_es_siem,
        "has_splunk": has_splunk,
        "cis_controls": ", ".join(sorted(cis_set))[:500],
        "nist_controls": ", ".join(sorted(nist_set))[:500],
        "process_coverage": False,
        "network_coverage": False,
        "file_coverage": False,
        "cloud_coverage": False,
        "hardware_coverage": False,
        "combined_score": round(combined, 4),
        "mitigation_score": round(mit_score, 4),
        "detection_score": round(det_score, 4),
        "weight": weight,
        "detection_level": detection_level,
        "ndi_score": ndi_score,
    }

def build_new_calculator(tid_map, mitigations, tech_mitigations, old_methodology, stix, id_to_tid, external_counts=None):
    """Build the new Calculator.xlsx with v19 data + preserved methodology + external detection counts."""
    wb = Workbook()
    
    # ===== TECHNIQUES SHEET =====
    ws_tech = wb.active
    ws_tech.title = "techniques"
    
    # Header
    headers = [
        "TID", "Name", "Description", "URL", "", "", "", "", 
        "Detection", "Platforms", "Data Sources", "Is Subtechnique", "Supertechnique"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws_tech.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    
    # Sort: parent techniques first, then sub-techniques
    parents = sorted([t for t in tid_map.values() if not t["is_subtechnique"] and not t["revoked"] and not t["deprecated"]],
                     key=lambda x: x["tid"])
    subtechniques = sorted([t for t in tid_map.values() if t["is_subtechnique"] and not t["revoked"] and not t["deprecated"]],
                           key=lambda x: x["tid"])
    
    row = 2
    added_count = 0
    for tech in parents + subtechniques:
        tid = tech["tid"]
        parent_tid = tid.split(".")[0] if "." in tid else None
        
        ws_tech.cell(row=row, column=1, value=tid)
        ws_tech.cell(row=row, column=2, value=RENAME_MAP.get(tid, tech["name"]))
        ws_tech.cell(row=row, column=3, value=tech["description"])
        ws_tech.cell(row=row, column=4, value=tech["url"])
        ws_tech.cell(row=row, column=9, value=tech["detection"])
        ws_tech.cell(row=row, column=10, value=", ".join(tech["platforms"]))
        ws_tech.cell(row=row, column=11, value=", ".join(tech["data_sources"]))
        ws_tech.cell(row=row, column=12, value=tech["is_subtechnique"])
        ws_tech.cell(row=row, column=13, value=parent_tid or "")
        
        # Add mitigations
        if tid in tech_mitigations:
            tech["mitigations"] = tech_mitigations[tid]
        
        row += 1
        added_count += 1
    
    print(f"Added {added_count} techniques to techniques sheet")
    
    # ===== MITIGATIONS SHEET =====
    ws_mit = wb.create_sheet("mitigations")
    mit_headers = ["MID", "", "Name", "Description", "URL"]
    for col, h in enumerate(mit_headers, 1):
        cell = ws_mit.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    
    row = 2
    for mid in sorted(mitigations.keys()):
        m = mitigations[mid]
        if m["deprecated"]:
            continue
        ws_mit.cell(row=row, column=1, value=mid)
        ws_mit.cell(row=row, column=3, value=m["name"])
        ws_mit.cell(row=row, column=4, value=m["description"])
        ws_mit.cell(row=row, column=5, value=m["url"])
        row += 1
    
    print(f"Added {row - 2} mitigations to mitigations sheet")
    
    # ===== RELATIONSHIPS SHEET =====
    ws_rel = wb.create_sheet("relationships")
    rel_headers = ["Source (Mitigation)", "", "", "", "", "Target (Technique)"]
    for col, h in enumerate(rel_headers, 1):
        cell = ws_rel.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    
    row = 2
    for tid in sorted(tech_mitigations.keys()):
        # Skip revoked/deprecated techniques
        if tid in tid_map and (tid_map[tid]["revoked"] or tid_map[tid]["deprecated"]):
            continue
        for mid in tech_mitigations[tid]:
            if mid in mitigations and not mitigations[mid]["deprecated"]:
                ws_rel.cell(row=row, column=1, value=mid)
                ws_rel.cell(row=row, column=6, value=tid)
                row += 1
    
    print(f"Added {row - 2} relationships to relationships sheet")
    
    # ===== METHODOLOGY SHEET =====
    ws_method = wb.create_sheet("Methodology")
    
    # Build merged methodology: old scores + defaults for new techniques
    merged_methodology = {}
    
    # First, carry over all old methodology scores (for techniques that still exist)
    mapped_scores = {}  # Track which old TIDs we've mapped to new TIDs
    for old_tid, scores in old_methodology.items():
        # Check if this TID still exists in v19
        if old_tid in tid_map and not tid_map[old_tid]["revoked"] and not tid_map[old_tid]["deprecated"]:
            merged_methodology[old_tid] = scores
            mapped_scores[old_tid] = True
        # Check revocation map
        elif old_tid in REVOCATION_MAP:
            new_tid = REVOCATION_MAP[old_tid]
            if new_tid not in merged_methodology:
                merged_methodology[new_tid] = scores
                mapped_scores[old_tid] = True
                print(f"  Mapped methodology: {old_tid} -> {new_tid}")
    
    # Count detections for CTID scoring
    detection_counts, analytic_counts = count_detections(stix, id_to_tid)
    
    # Build MID -> CIS/NIST mapping from existing scored techniques
    mid_to_cis, mid_to_nist = build_mid_control_mapping(old_methodology, tech_mitigations)
    
    # For new v19 techniques without scores, compute CTID methodology scores
    new_count = 0
    for tid in sorted(tid_map.keys()):
        t = tid_map[tid]
        if t["revoked"] or t["deprecated"]:
            continue
        if tid not in merged_methodology:
            merged_methodology[tid] = compute_ctid_defaults(
                tid, tid_map, stix, id_to_tid, detection_counts,
                tech_mitigations, mid_to_cis, mid_to_nist, old_methodology,
                external_counts=external_counts
            )
            new_count += 1
    
    print(f"Methodology: {len(merged_methodology)} total ({len(mapped_scores)} preserved, {new_count} CTID-scored)")
    
    # ===== NDI FIELD COMPUTATION: Add weight (W), detection_level (D), ndi_score for ALL techniques =====
    ndi_entries = []
    for tid in sorted(merged_methodology.keys()):
        scores = merged_methodology[tid]
        # Get tactics from tid_map
        tactics = tid_map.get(tid, {}).get("tactics", []) if tid in tid_map else []
        weight = compute_weight(tactics)
        
        # Get existing detection_score from methodology
        det_score = scores.get("detection_score", 0) or 0
        stix_det = detection_counts.get(tid, 0)
        
        detection_level = compute_detection_level(tid, external_counts, float(det_score), stix_det)
        ndi_score = weight * detection_level
        
        scores["weight"] = weight
        scores["detection_level"] = detection_level
        scores["ndi_score"] = ndi_score
        ndi_entries.append((weight, detection_level))
    
    # Compute overall NDI for the full technique set
    overall_ndi, num, den = compute_ndi_scores(ndi_entries)
    print(f"NDI: overall={overall_ndi}%, numerator={num}, denominator={den} ({len(ndi_entries)} techniques)")
    
    # ===== SURGICAL RECALC: Only fix the 141 flat detection_score=0.01 techniques =====
    # Load original scores from Techniques.json as ground truth for non-flat techniques
    TECHNIQUES_JSON_PATH = os.path.join(REPO_DIR, "src", "data", "Techniques.json")
    original_scores = {}  # tid -> {detection_score, combined_score, cumulative_score}
    flat_01_tids = set()
    if os.path.exists(TECHNIQUES_JSON_PATH):
        with open(TECHNIQUES_JSON_PATH, encoding="utf-8") as f:
            orig_techs = json.load(f)
        for t in orig_techs:
            tid = t.get("tid", "")
            ds = t.get("actionability_score", {}).get("detection_score")
            cs = t.get("actionability_score", {}).get("combined_score")
            cum = t.get("cumulative_score")
            original_scores[tid] = {
                "detection_score": ds,  # may be None
                "combined_score": cs,
                "cumulative_score": cum,
            }
            if ds is not None and abs(float(ds) - 0.01) < 0.001:
                flat_01_tids.add(tid)
        n_none = sum(1 for v in original_scores.values() if v["detection_score"] is None)
        print(f"Original Techniques.json: {len(original_scores)} loaded "
              f"({len(flat_01_tids)} flat 0.01, {n_none} null)")
    
    for tid, scores in merged_methodology.items():
        orig = original_scores.get(tid, {})
        orig_ds = orig.get("detection_score")
        
        if tid in flat_01_tids:
            # This technique had flat 0.01 — try to differentiate with external data
            ext = external_counts.get(tid, {}) if external_counts else {}
            total_det = ext.get("total_detections", 0)
            sightings = ext.get("sightings", 0)
            
            if total_det > 0 or sightings > 0:
                stix_det = detection_counts.get(tid, 0)
                new_det_score = compute_detection_score_from_external(tid, external_counts or {}, stix_det)
                
                old_mit_score = float(scores.get("mitigation_score", 0) or 0)
                
                # Recompute combined_score
                w_m = 1.0
                w_d = 0.5 * (DET_UPPER - DET_LOWER) / (MIT_UPPER - MIT_LOWER)
                total_w = w_d + w_m
                w_d_norm = w_d / total_w
                w_m_norm = w_m / total_w
                new_combined = w_d_norm * new_det_score + w_m_norm * old_mit_score
                
                scores["detection_score"] = round(new_det_score, 4)
                scores["combined_score"] = round(new_combined, 4)
                scores["cumulative_score"] = round(new_combined, 4)
                
                # Update has_* booleans from real counts
                scores["has_car"] = ext.get("car", 0) > 0
                scores["has_sigma"] = ext.get("sigma", 0) > 0
                scores["has_es_siem"] = ext.get("es", 0) > 0
                scores["has_splunk"] = ext.get("splunk", 0) > 0
                
                # Recompute detection_level and ndi_score with new data
                tactics = tid_map.get(tid, {}).get("tactics", []) if tid in tid_map else []
                weight = scores.get("weight", compute_weight(tactics))
                new_det_level = compute_detection_level(tid, external_counts or {}, new_det_score, stix_det)
                scores["weight"] = weight
                scores["detection_level"] = new_det_level
                scores["ndi_score"] = weight * new_det_level
            else:
                # No external data — use STIX fallback for at least SOME differentiation
                stix_det = detection_counts.get(tid, 0)
                new_det_score = compute_detection_score_from_external(tid, external_counts or {}, stix_det)
                
                old_mit_score = float(scores.get("mitigation_score", 0) or 0)
                w_m = 1.0
                w_d = 0.5 * (DET_UPPER - DET_LOWER) / (MIT_UPPER - MIT_LOWER)
                total_w = w_d + w_m
                w_d_norm = w_d / total_w
                w_m_norm = w_m / total_w
                new_combined = w_d_norm * new_det_score + w_m_norm * old_mit_score
                
                scores["detection_score"] = round(new_det_score, 4)
                scores["combined_score"] = round(new_combined, 4)
                scores["cumulative_score"] = round(new_combined, 4)
                
                # Recompute detection_level and ndi_score
                tactics = tid_map.get(tid, {}).get("tactics", []) if tid in tid_map else []
                weight = scores.get("weight", compute_weight(tactics))
                new_det_level = compute_detection_level(tid, external_counts or {}, new_det_score, stix_det)
                scores["weight"] = weight
                scores["detection_level"] = new_det_level
                scores["ndi_score"] = weight * new_det_level
        elif orig:
            # Restore original scores from Techniques.json (ground truth)
            scores["detection_score"] = orig_ds  # may be None
            if orig.get("combined_score") is not None:
                scores["combined_score"] = orig["combined_score"]
            if orig.get("cumulative_score") is not None:
                scores["cumulative_score"] = orig["cumulative_score"]
    
    # Count flat 0.01 remaining
    flat_after = sum(1 for s in merged_methodology.values() 
                     if s.get("detection_score") is not None and abs(float(s.get("detection_score", 0)) - 0.01) < 0.001)
    n_differentiated = len(flat_01_tids) - flat_after if flat_01_tids else 0
    print(f"Surgical recalc: {n_differentiated} flat techniques differentiated")
    print(f"  Flat 0.01 before: {len(flat_01_tids)}, after: {flat_after}")
    
    # Write Methodology sheet
    # Column layout matching update_techniques.js expectations:
    # B=cumulative_score, C=TID, H=choke_point(8), M=prevalence(13)
    # N=has_car, O=has_sigma, P=has_es_siem, Q=has_splunk
    # R=cis_controls, T=nist_controls
    # V=combined(22), Y=mitigation(25), AB=detection(28)
    # AE=process(31), AG=network(33), AI=file(35), AK=cloud(37), AM=hardware(39)
    # AN=weight(40), AP=detection_level(42), AR=ndi_score(44)
    
    row = 1
    for tid in sorted(merged_methodology.keys()):
        s = merged_methodology[tid]
        row += 1
        ws_method.cell(row=row, column=1, value=s.get("cumulative_score"))  # A - could also be cum score
        ws_method.cell(row=row, column=2, value=s.get("cumulative_score"))  # B
        ws_method.cell(row=row, column=3, value=tid)  # C - Technique ID
        ws_method.cell(row=row, column=8, value=s.get("choke_point_score"))  # H
        ws_method.cell(row=row, column=13, value=s.get("prevalence_score"))  # M
        ws_method.cell(row=row, column=14, value=1 if s.get("has_car") else 0)  # N
        ws_method.cell(row=row, column=15, value=1 if s.get("has_sigma") else 0)  # O
        ws_method.cell(row=row, column=16, value=1 if s.get("has_es_siem") else 0)  # P
        ws_method.cell(row=row, column=17, value=1 if s.get("has_splunk") else 0)  # Q
        ws_method.cell(row=row, column=18, value=s.get("cis_controls", ""))  # R
        ws_method.cell(row=row, column=20, value=s.get("nist_controls", ""))  # T
        ws_method.cell(row=row, column=22, value=s.get("combined_score"))  # V
        ws_method.cell(row=row, column=25, value=s.get("mitigation_score"))  # Y
        ws_method.cell(row=row, column=28, value=s.get("detection_score"))  # AB
        ws_method.cell(row=row, column=31, value=1 if s.get("process_coverage") else 0)  # AE
        ws_method.cell(row=row, column=33, value=1 if s.get("network_coverage") else 0)  # AG
        ws_method.cell(row=row, column=35, value=1 if s.get("file_coverage") else 0)  # AI
        ws_method.cell(row=row, column=37, value=1 if s.get("cloud_coverage") else 0)  # AK
        ws_method.cell(row=row, column=39, value=1 if s.get("hardware_coverage") else 0)  # AM
        ws_method.cell(row=row, column=40, value=s.get("weight"))  # AN - NDI Weight (W)
        ws_method.cell(row=row, column=42, value=s.get("detection_level"))  # AP - NDI Detection Level (D)
        ws_method.cell(row=row, column=44, value=s.get("ndi_score"))  # AR - NDI Score (W×D)
    
    # Save
    wb.save(NEW_CALC_PATH)
    print(f"\nSaved new Calculator.xlsx to {NEW_CALC_PATH}")

def main():
    print("Loading STIX v19.1 data...")
    stix = load_stix()
    
    print("Extracting techniques...")
    tid_map, id_to_tid = extract_techniques(stix)
    
    active = [t for t in tid_map.values() if not t["revoked"] and not t["deprecated"]]
    parents = [t for t in active if not t["is_subtechnique"]]
    subs = [t for t in active if t["is_subtechnique"]]
    print(f"  Active: {len(parents)} parent + {len(subs)} sub-techniques = {len(active)} total")
    print(f"  Revoked/deprecated: {sum(1 for t in tid_map.values() if t['revoked'] or t['deprecated'])}")
    
    print("Extracting mitigations...")
    mitigations = extract_mitigations(stix)
    print(f"  {len(mitigations)} mitigations")
    
    print("Extracting relationships...")
    tech_mitigations = extract_relationships(stix, id_to_tid, mitigations)
    total_rels = sum(len(v) for v in tech_mitigations.values())
    print(f"  {total_rels} mitigation-to-technique relationships")
    
    print("Loading existing methodology...")
    old_methodology = load_existing_methodology()
    
    print("Loading external detection counts...")
    external_counts = load_external_counts()
    
    print("Building new Calculator.xlsx with CTID scoring...")
    build_new_calculator(tid_map, mitigations, tech_mitigations, old_methodology, stix, id_to_tid, external_counts=external_counts)

if __name__ == "__main__":
    main()
